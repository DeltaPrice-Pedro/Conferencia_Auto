from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilenames, askopenfilename, asksaveasfilename
import tabula as tb
import pandas as pd

from unidecode import unidecode
import string
import os
import copy
import sys
from datetime import *
import locale
import re

from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

thin_border = Border(
    left=Side(border_style='thin', color='FF000000'),
    right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'),
    bottom=Side(border_style='thin', color='FF000000')
)

dashed_border = Border(
    left=Side(border_style='dashed', color='FF000000'),
    right=Side(border_style='dashed', color='FF000000'),
    top=Side(border_style='dashed', color='FF000000'),
    bottom=Side(border_style='dashed', color='FF000000')
)

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

window = Tk()

class Arquivo:
    def __init__(self):
        self.tipos_validos = ''

    def validar_entrada(self, caminho):
        if any(c not in string.ascii_letters for c in caminho):
            caminho = self.formato_ascii(caminho)

        self.__tipo(caminho)
        return caminho

    def __tipo(self, caminho):
        if caminho[len(caminho) -3 :] != self.tipos_validos:
            ultima_barra = caminho.rfind('/')
            raise Exception(
                f'Formato inválido do arquivo: {caminho[ultima_barra+1:]}')

    def formato_ascii(self, caminho):
        caminho_uni = unidecode(caminho)
        os.renames(caminho, caminho_uni)
        return caminho_uni
    
    def envio_invalido(self):
        return True if len(self.caminho) == 0 else False

class Matriz(Arquivo):
    def __init__(self):
        super().__init__()
        self.tipos_validos = 'lsx'
        self.caminho = ''

    def inserir(self, label):
        try:
            caminho = askopenfilename()

            if caminho == '':
                return None

            caminho_validado = self.validar_entrada(caminho)
            label['text'] = caminho[caminho.rfind('/') +1:]

            self.caminho = caminho_validado

        except PermissionError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado apresenta-se aberto em outra janela, favor fecha-la')
        except FileExistsError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado já apresenta uma versão sem acento, favor usar tal versão ou apagar uma delas')
        except Exception as error:
            messagebox.showerror(title='Aviso', message= error)

    def ler(self):
        return pd.read_excel(self.caminho, na_filter=False, usecols='A:B')
    
    def load(self):
        return load_workbook(self.caminho)
    
class Recibo(Arquivo):
    def __init__(self):
        super().__init__()
        self.tipos_validos = 'pdf'
        self.caminho = []

    def get_caminho(self):
        return self.caminho

    def inserir(self, label):
        try:
            caminhos = askopenfilenames()

            if caminhos == '':
                return None

            label.delete(0, END)
            caminhos_validados = []
            for item in caminhos:
                caminhos_validados.append(self.validar_entrada(item))
                label.insert(END, f'{item[item.rfind('/') +1:]}\n')

            self.caminho = caminhos_validados.copy()

        except PermissionError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado apresenta-se em aberto em outra janela, favor fecha-la')
        except FileExistsError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado já apresenta uma versão sem acento, favor usar tal versão ou apagar uma delas')
        except Exception as error:
            messagebox.showerror(title='Aviso', message= error)

class IFielding:

    def field(self, df_recibo, ws):
        for col_index, columns in enumerate(list(df_recibo.columns), 1):
            celula = ws.cell(self.LIN_INDEX, col_index, columns)
            celula.font = Font(bold=True, underline='single')
            celula.alignment = Alignment(horizontal='center')
            celula.border = Border(top=Side(border_style='medium', color='FF000000'))
            celula.fill = PatternFill(start_color='B0BFB2',
                                        end_color='B0BFB2',
                                        fill_type='solid')

class IDating:

    def data(self, df_matriz, df_recibo, data_confe):
        excluido = Adcional('FFF500', df_recibo, f'NÃO RELACIONADAS {self.titulo}')
        repetido = Adcional('7795FF', df_recibo, f'REPETIDAS {self.titulo}')
        atrasado = Adcional('FF6563', df_recibo, f'FORA DA COMPETÊNCIA {self.titulo}')

        adcionais = [excluido, repetido, atrasado]
        for index_recibo, row_recibo in df_recibo.iterrows():
            #print(f'{row_recibo} - CNPJ procurado')
            for index_matriz, row_matriz in df_matriz.iterrows():
                #print(f'{row_matriz} - opções')
                if row_recibo['Referência'] != data_confe:
                    atrasado.add_data(row_recibo)
                    break
                if row_recibo['CNPJ'] == row_matriz['CNPJ']:
                    if self.ws.cell(index_matriz + self.LIN_DATA, 3).value != '':
                        repetido.add_data(row_recibo)
                    else:
                        for col_index, valor in enumerate(row_recibo, 1):
                            celula = self.ws.cell\
                            (index_matriz + self.LIN_DATA, col_index, valor)
                            celula.alignment = Alignment(horizontal='center')
                            celula.border = dashed_border
                    break
                excluido.add_data(row_recibo)
        
        return adcionais
    
    def width_ws(self, ws):
        for index, valor in enumerate([40,20,15,20,15,20,20],1):
            ws.column_dimensions[get_column_letter(index)].width = valor

    def valid_adcionais(self, adcionais):
        ref = ['Não Relacionados', 'Repetidos', 'Fora da Competência']

        for index, adc in enumerate(adcionais):
            if adc.qnt_data() != 0:
                if ref[index] in self.wb.sheetnames:
                    ws = self.wb[ref[index]]
                else:
                    ws = self.wb.create_sheet(ref[index])
                self.width_ws(ws)
                adc.preencher(ws)
        
class Criacao (IFielding, IDating):
    def __init__(self, titulo):
        self.LIN_INDEX = 7
        self.LIN_DATA = 8
        self.titulo = titulo

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Relacionados'
        self.width_ws(self.ws)

    def criar(self, df_matriz, df_recibo, nome_arq, data_confe):
        self._cabecalho(data_confe)
        self._table_ref(df_matriz)
        self.field(df_recibo, self.ws)
        self._matriz(df_matriz, df_recibo)
        
        adcionais = self.data(df_matriz, df_recibo, data_confe)

        self.valid_adcionais(adcionais)

        self.wb.save(nome_arq+'.xlsx') 

        return adcionais 

    def _cabecalho(self, data_confe):  
        self.ws.cell(1,1, f'RELATÓRIO DE CONFERÊNCIA {self.titulo}').font = Font(size=26,
                bold=True,)
        
        celula = self.ws.cell(3,1, 'Competência')
        celula.font = Font(size=16, bold=True)
        celula.alignment = Alignment(horizontal= 'right')

        self.ws.cell(3,2, data_confe)
        
        celula = self.ws.cell(4,1, 'Data Entrega')
        celula.font = Font(size=16, bold=True)
        celula.alignment = Alignment(horizontal= 'right')

        self.ws.cell(4,2, datetime.now().strftime("%d/%m/%Y"))

    def _table_ref(self, df_matriz):
        tam_df = len(df_matriz)+7
        ref = {
            'Obrigadas:': f'=COUNTA($A8:$A{tam_df})',
            'Entregues:': f'=COUNTA($C8:$C{tam_df})',
            'Não Entregues:': '=$E3 - $E4'
        }

        #f'=COUNTIF($D8:$D{tam_df}, "ENVIADO")'

        for index, text in enumerate(ref.items(), 1):
            celula = self.ws.cell(index + 2, 4, text[0])
            celula.font = Font(bold=True)
            celula.alignment = Alignment(horizontal= 'right')
            celula.border = thin_border
            
            celula = self.ws.cell(index + 2, 5, text[1])
            celula.alignment = Alignment(horizontal= 'center')
            celula.border = thin_border
            
    def _matriz(self, df_matriz, df_recibo):
        for index, row in df_matriz.iterrows():
            #Adciona Nome e CNPJ apenas
            for col_index, valor in enumerate(row, 1):
                celula = self.ws.cell(index + self.LIN_DATA, col_index, valor)
                celula.alignment = Alignment(horizontal= 'center')
                celula.border = dashed_border
            
            #Termina o df com espaços vazios
            self._espacos_vazios(index, df_recibo)

    def _espacos_vazios(self, index, df_recibo):
        dif_cnpj = 2
        for col_index in range(len(df_recibo.columns) - dif_cnpj):
            self.ws.cell(index + self.LIN_DATA, col_index + dif_cnpj + 1, '')\
                .border = dashed_border

class Incremento (IDating):
    def __init__(self, wb, titulo):
        self.LIN_DATA = 9
        self.titulo = titulo
        
        self.wb = wb
        self.ws = self.wb['Relacionados']

    def incrementar(self, df_matriz, df_relatorio, nome_arq):
        data_relatorio = self.ws.cell(3,2).value

        adcionais = self.data(
            self._init_matriz(df_matriz), df_relatorio, data_relatorio)

        self.valid_adcionais(adcionais)

        self.wb.save(nome_arq+'.xlsx')

        return adcionais

    def _init_matriz(self, df_matriz):
        titulo_arq = df_matriz.columns[0]
        if self.titulo not in titulo_arq:
            raise Exception('O relatório inserido é de uma obrigação diferente do recibo em questão')

        df_matriz.columns = ["EMPRESA","CNPJ"]
        return df_matriz.drop([0,1,2,3,4,5,6])\
            .reset_index(drop=True)

class Adcional(IFielding):
    def __init__(self, cor, df_recibo, titulo):
        self.data = []
        self.LIN_INDEX = 3
        self.titulo = titulo
        self.df_recibo = df_recibo
        self.cor = cor

    def add_data(self, row):
        self.data.append(row)

    def qnt_data(self):
        return len(self.data)

    def preencher(self, ws):
        self._titulo(ws)
        self.field(self.df_recibo, ws)
        self._data(ws)

    def _titulo(self, ws):
        ws.cell(1,1, self.titulo).font = Font(size=26, bold=True,)

    def _data(self, ws):
        for index_recibo, row_recibo in enumerate(self.data, 1):
            index_recibo = self._lin_disp(index_recibo, ws)
            for col_index, valor in enumerate(row_recibo, 1):
                celula = ws.cell(index_recibo + self.LIN_INDEX, col_index, valor)
                celula.alignment = Alignment(horizontal='center')
                celula.border = dashed_border
                celula.fill = PatternFill(start_color= self.cor,
                                        end_color= self.cor,
                                        fill_type='solid')

    def _lin_disp(self, index_recibo, ws):
        if ws.cell(index_recibo + self.LIN_INDEX, 1).value == None:
            return index_recibo
        return self._lin_disp(index_recibo + 1, ws)

class Competencia:
    def __init__(self):
        self.nome_emp = []
        self.cnpj = []
        self.referencia = []
        self.data = []
        self.hora = []

    def to_string(self):
        return self.titulo

class Des(Competencia):
    def __init__(self):
        super().__init__()
        self.servicos = []
        self.titulo = 'DES'

    def add_linha(self, arquivo):
        tabela = tb.read_pdf(arquivo, pages= 1, stream= True,\
                        relative_area=True, area= [10,0,59,68])[0]
        
        tabela.fillna('', inplace=True)

        ##Nome Emp
        self.nome_emp.append(tabela.iloc[1,0].replace('Nome/Razão Social: ',''))

        ##CNPJ
        if tabela.iloc[0,1] == '':
            self.cnpj.append(tabela.iloc[0,0][45:])
        else:
            self.cnpj.append(tabela.iloc[0,1])

        ##Ref.
        data_ref = tabela.iloc[3,0].replace('Referência: ','').replace(' No Protocolo:','')
        data_format = datetime.strptime(data_ref, '%B/%Y')

        self.referencia.append(data_format.strftime("%m/%Y"))

        ##Data e Hora
        col_dthr = tabela.iloc[4,0].replace('Data/Hora de Entrega: ','')\
                        .replace(' Regime de Tributação:','')

        self.data.append(col_dthr[:10])

        self.hora.append(col_dthr[10:])

        ##Serviços declarados.
        self.servicos.append(tabela.iloc[15,0]\
            .replace('Total de Serviços Declarados: ','')\
                .replace('Base de Cálculo S/ Ret',''))

    def gerar_df(self):
        return pd.DataFrame({
            'Nome Empresa': self.nome_emp,
            'CNPJ': self.cnpj, 
            'Referência': self.referencia, 
            'Data Entrega': self.data,
            'Hora Entrega': self.hora,
            'Serviços Tomados': self.servicos
            })

class Reinf(Competencia):
    def __init__(self):
        super().__init__()
        self.num_dom = []
        self.situacao = []
        self.titulo = 'EFD REINF'

    def add_linha(self, arquivo):
        arquivo = tb.read_pdf(arquivo, pages=1, stream=True,\
                        relative_area=True ,area=[5,0,100,100])[0]

        tabela = arquivo.loc[arquivo['Unnamed: 2'] == 'R-2099 - Fechamento dos Eventos Periódicos']

        prim_linha = tabela.iloc[0,0]

        ##Num Domínio
        self.num_dom.append(prim_linha[:prim_linha.find('-')-1])

        ##Nome empresa
        self.nome_emp.append(prim_linha[prim_linha.find('-')+2:])

        ##CNPJ
        self.cnpj.append(tabela.iloc[0,1][:18])

        ##Ref
        self.referencia.append(tabela.iloc[0,2])

        ##Situação
        self.situacao.append(tabela.iloc[0,5].replace('Sucesso','ENVIADO'))

        ##Data e Hora
        col_dthr = tabela.iloc[0,7][:18]

        self.data.append(col_dthr[:10])

        self.hora.append(col_dthr[12:])

    def gerar_df(self):
        return pd.DataFrame({
            'Nome Empresa': self.nome_emp,
            'CNPJ': self.cnpj, 
            'Num. Domínio': self.num_dom,
            'Referência': self.referencia,
            'Situação': self.situacao, 
            'Data Entrega': self.data,
            'Hora Entrega': self.hora,
            })
    
class Contribuicoes(Competencia):
    def __init__(self):
        super().__init__()
        self.titulo = 'EFD CONTRIBUIÇÕES'

    def add_linha(self, arquivo):
        tabela = tb.read_pdf(arquivo, pages=1, stream=True,\
                        relative_area=True ,area=[10,0,100,100])[0]

        dif = 0
        dif_dthr = 0

        if 'IDENTIFICAÇÃO' not in tabela.iloc[0,0]:
            dif = 1
            dif_dthr = 20

        ##Nome Empresa
        self.nome_emp.append(tabela.iloc[1 + dif,0].replace('Contribuinte: ',''))

        ##CNPJ
        self.cnpj.append(tabela.iloc[2 + dif,0][6:24])

        ##Ref
        self.referencia.append(tabela.iloc[4 + dif,0][37:])

        ##Data e Hora
        col_dthr = tabela.iloc[28,0]

        self.data.append(col_dthr[3 + dif_dthr :13 + dif_dthr])

        self.hora.append(col_dthr[17 + dif_dthr:])

    def gerar_df(self):
        return pd.DataFrame({
            'Nome Empresa': self.nome_emp,
            'CNPJ': self.cnpj, 
            'Referência': self.referencia,
            'Data Entrega': self.data,
            'Hora Entrega': self.hora,
            })
    
class SimplesNacional(Competencia):
    def __init__(self):
        super().__init__()
        self.valor = []
        self.anexo = []
        self.titulo = 'SIMPLES NACIONAL'

    def add_linha(self, arquivo):
        tabela = tb.read_pdf(arquivo, pages=1, stream=True,\
                        relative_area=True ,area=[0.5,0,100,100])[0]

        ##Nome Empresa
        self.nome_emp.append(tabela.iloc[18,0].replace('Estabelecimento: ','')[3:])

        ##CNPJ
        self.cnpj.append(tabela.iloc[0,0].replace('CNPJ: ',''))

        ##Ref
        self.referencia.append(tabela.iloc[4,0].replace('Período: ',''))

        ##Data
        self.data.append(str(tabela.iloc[0,1]).replace('Emissão: ',''))

        ##Valor Tributo
        self.valor.append(tabela.iloc[len(tabela)-2,0].replace('Simples Nacional a recolher: ',''))

        ##Anexo
        self.anexo.append(tabela.iloc[19,0][7:15].strip())

    def gerar_df(self):
        return pd.DataFrame({
            'Nome Empresa': self.nome_emp,
            'CNPJ': self.cnpj, 
            'Referência': self.referencia,
            'Data Entrega': self.data,
            'Valor Tributo': self.valor,
            'Anexo': self.anexo
            })

class DCTF(Competencia):
    def __init__(self):
        super().__init__()
        self.valor = []
        self.titulo = 'DCTF'

    def sum_saldos(self, arquivo):
        valor_valido = []
        for row in arquivo.iloc[10:21,0]:
            pos_vir = row.find(',')
            valor = row[pos_vir+3:].replace('.','').replace(',','.').strip()
            if valor[:4] != '0.00':
                valor_valido.append(float(valor))
        if len(valor_valido) != 0:
            return sum(valor_valido)
        return 0

    def add_linha(self, arquivo):
        tabela = tb.read_pdf(arquivo, pages=1, stream=True, guess=False)[0]

        ##Nome
        self.nome_emp.append(tabela.iloc[4,0].replace('Nome Empresarial: ',''))

        ##CNPJ
        self.cnpj.append(tabela.iloc[3,0][6:24])

        ##Ref

        data_ref = tabela.iloc[3,0][34:]
        data_format = datetime.strptime(data_ref, '%b %Y')

        self.referencia.append(data_format.strftime("%m/%Y"))

        data_row = tabela.iloc[54,0]\
            .replace('exigido este número de recibo: em ','')

        ##Data
        self.data.append(data_row[:10])

        ##Hora
        self.hora.append(data_row[13:])

        ##Valor Tributo
        self.valor.append(f'{self.sum_saldos(tabela):.2f}')

    def gerar_df(self):
        return pd.DataFrame({
            'Nome Empresa': self.nome_emp,
            'CNPJ': self.cnpj,
            'Referência': self.referencia,
            'Data': self.data,
            'Hora': self.hora,
            'Saldo em Aberto': self.valor
            })

class App:
    def __init__(self):
        self.window = window
        self.recibos = Recibo()
        self.matriz = Matriz()

        self.ref = {
            'des' : Des(),
            'reinf' : Reinf(),
            'contribuicoes': Contribuicoes(),
            'contribuições' : Contribuicoes(),
            'simples nacional': SimplesNacional(),
            'sn': SimplesNacional(),
            'dctf': DCTF()
        }

        self.tela()
        self.main()
        window.mainloop()

    def tela(self):
        self.window.configure(background='darkblue')
        self.window.resizable(False,False)
        self.window.geometry('860x500')
        self.window.iconbitmap(self.resource_path('imgs\\conf-icon.ico'))
        self.window.title('Conversor de Extrato')

    def resource_path(self,relative_path):
        base_path = getattr(
            sys,
            '_MEIPASS',
            os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)

    def comp_formater(self, text, var, index, mode): 
        #Só recebe valor que passa pelo validador
        valor = text.get()
        if len(valor) == 6 and '/' not in valor:
           valor = valor[:2] + "/" + valor[2:]
        else:
            valor = valor.replace('/','')
        text.set(valor)

    def comp_validator(self, text):
        padrao = r"^[-\d.,/]+$"  # Permite dígitos, ponto, vírgula, hífen e barra
        if len(text) < 8:
            if len(text) >= 7:
                return re.match(padrao, text) is not None
            elif len(text) in [0,6] or text.isdecimal():
                return True
        return False

    def main(self):
        self.index = Frame(self.window, bd=4, bg='lightblue')
        self.index.place(relx=0.05,rely=0.05,relwidth=0.9,relheight=0.9)

        #Titulo
        Label(self.index, text='Conferência Automática', background='lightblue', font=('arial',30,'bold')).place(relx=0.23,rely=0.18,relheight=0.15)

        #Logo
        self.logo = PhotoImage(file=self.resource_path('imgs\\conferencia_horizontal.png'))
        
        self.logo = self.logo.subsample(3,3)
        
        Label(self.window, image=self.logo, background='lightblue', border=0)\
            .place(relx=0.175,rely=0.05,relwidth=0.7,relheight=0.2)

        self.valIncrement = BooleanVar()

        self.valIncrement.set(False)

        Radiobutton(self.index, text="Criar novo Relatório", value=False, variable= self.valIncrement, command= lambda: self.entryCompe.config(state='normal')).place(relx=0.45,rely=0.33)

        Radiobutton(self.index, text="Incrementar em Relatório antigo", value=True, variable= self.valIncrement, command= lambda: self.entryCompe.config(state='disabled')).place(relx=0.65,rely=0.33)

        #Labels e Entrys
        ###########Matriz
        Label(self.index, text='Insira aqui a Matriz/Referência:',\
            background='lightblue', font=(10))\
                .place(relx=0.15,rely=0.33)

        self.matLabel = Label(self.index)
        self.matLabel.config(font=("Arial", 12, 'bold italic'), anchor= 's')
        self.matLabel.place(relx=0.21,rely=0.4,relwidth=0.7, relheight=0.055)
        
        Button(self.index, text='Enviar',\
            command = lambda: self.matriz.inserir(self.matLabel))\
                .place(relx=0.15,rely=0.4,relwidth=0.06,relheight=0.055)

        ###########Arquivo
        Label(self.index, text='Insira aqui os Recibos:',\
            background='lightblue', font=(10))\
                .place(relx=0.15,rely=0.48)

        self.arqLabel = Listbox(self.index, border= 0)
        self.arqLabel.config(font=("Arial", 8, 'bold italic'))
        self.arqLabel.place(relx=0.21,rely=0.55,relwidth=0.675, relheight=0.2)

        self.barra = Scrollbar(self.index, command= self.arqLabel.yview)\
            .place(relx=0.875,rely=0.55,relwidth=0.03, relheight=0.2)
        
        self.arqLabel.config(yscrollcommand= self.barra)
        
        Button(self.index, text='Enviar',\
            command = lambda: self.recibos.inserir(self.arqLabel))\
                .place(relx=0.15,rely=0.55,relwidth=0.06,relheight=0.055)

        ###########EFD
        Label(self.index, text='Caso o nome da obrigação assesória não constar no nome do arquivo',\
            background='lightblue', font=("Arial", 12, 'bold italic'))\
                .place(relx=0.2,rely=0.775)

        Label(self.index, text='Escolha a obrigação:',\
            background='lightblue', font=(10))\
                .place(relx=0.2,rely=0.825)
        
        self.declaracaoEntry = StringVar()

        self.declaracaoEntryOpt = ["DES", "REINF", "CONTRIBUIÇÕES", "SIMPLES NACIONAL", "DCTF"]

        self.declaracaoEntry.set('Escolha aqui')

        self.popup = OptionMenu(self.index, self.declaracaoEntry, *self.declaracaoEntryOpt)\
            .place(relx=0.425,rely=0.835,relwidth=0.2,relheight=0.06)
        
        ###########Data Competência
        
        self.dt_compe = StringVar()

        self.dt_compe.trace_add('write', lambda *args, passed = self.dt_compe:\
            self.comp_formater(passed, *args) )

        Label(self.index, text='Data da Competência:',\
            background='lightblue', font=(10))\
                .place(relx=0.2,rely=0.925)
        

        self.entryCompe = Entry(self.index, textvariable = self.dt_compe, \
            validate ='key', validatecommand =(self.index.register(self.comp_validator), '%P'))
        
        self.entryCompe.place(relx=0.425,rely=0.925,relwidth=0.08,relheight=0.05)

        #Botão enviar
        Button(self.index, text='Gerar Conferencia',\
            command= lambda: self.executar())\
                .place(relx=0.65,rely=0.85,relwidth=0.25,relheight=0.12)
        
    def declaracao(self):
        itens_label = self.arqLabel.get(0,END)
        lista_declara = []

        primeiro_obj = copy.deepcopy(self.declaracao_valid(itens_label[0]))

        for itens in itens_label:
            lista_declara.append(type(self.declaracao_valid(itens)))

        if len(set(lista_declara)) != 1:
            raise Exception('Nem todos elementos são da mesma obrigação, favor selecionar tipo')

        return primeiro_obj
    
    def declaracao_valid(self, valor):
        if self.declaracaoEntry.get() != 'Escolha aqui':
            for key, obj in self.ref.items():
                if key in self.declaracaoEntry.get().lower():
                    return obj
        else:
            for chave, obj in self.ref.items():
                if chave in valor.lower():
                    return obj
            raise Exception('Nome da obrigação não identificado em todos os arquivos, favor selecionar tipo')
        
    def nomear_arq(self):
        nome_arq = asksaveasfilename(title='Favor selecionar a pasta onde será salvo', filetypes=((".xlsx","*.xlsx"),))

        if nome_arq == '':
            if messagebox.askyesno(title='Aviso', message= 'Deseja cancelar esta operação?') == True:
                raise Exception ('Operação cancelada!')
            else:
                return self.nomear_arq()
            
        return nome_arq  

    def avisar_adcionais(self, adcionais):
        ref = [
            'na aba "Não relacionadas" por não constarem na matriz',
            'com duplicidade. A segunda cópia foi inserida na aba "Repetidos"',
            'com data de competência desigual ao informado. Estas foram separadadas na aba "Fora da Competência"'
            ]

        for index, adc in enumerate(adcionais):
            if adc.qnt_data() != 0:
                messagebox.showinfo(title='Aviso', message= f'{adc.qnt_data()} empresas foram inseridas {ref[index]}') 

    def _validar_compe(self):
        if self.valIncrement.get() == False:
            datetime.strptime(self.dt_compe.get(), '%m/%Y')

    def executar(self):
        try:
            if self.matriz.envio_invalido():
                raise Exception ('Insira alguma Matriz')
            elif self.recibos.envio_invalido():
                raise Exception ('Insira algum Recibo')
            
            self._validar_compe()

            declaracao = self.declaracao()
            
            for arquivo in self.recibos.get_caminho():
                declaracao.add_linha(arquivo)

            df_recibo = declaracao.gerar_df()

            df_matriz = self.matriz.ler()

            nome_arq = self.nomear_arq()
            
            if self.valIncrement.get() == True:
                wb_completo = self.matriz.load()
                adcionais = Incremento(wb_completo, declaracao.to_string()).incrementar(df_matriz, df_recibo, nome_arq)
            else:
                adcionais = Criacao(declaracao.to_string()).criar(
                    df_matriz, df_recibo, nome_arq, self.dt_compe.get())

            self.avisar_adcionais(adcionais)

            messagebox.showinfo(title='Aviso', message='Abrindo o arquivo gerado!')

            os.startfile(nome_arq+'.xlsx')
         
        except (IndexError, TypeError):
            messagebox.showerror(title='Aviso', message= 'Erro ao extrair o recibo, confira se a obrigação foi selecionada corretamente. Caso contrário, comunique ao desenvolvedor')
        except KeyError:
            messagebox.showerror(title='Aviso', message= 'Relatório ou Matriz inserido é inválido, certifique-se que inseriu o documento correto')
        except ValueError:
            messagebox.showerror(title='Aviso', message= 'Data de Competência inserida é inválida')
        except Exception as error:
            messagebox.showerror(title='Aviso', message= error)
       
App()